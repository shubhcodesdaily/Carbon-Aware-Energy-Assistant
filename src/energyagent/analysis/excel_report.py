"""Export the analysis to a formatted Excel workbook: an hour-of-day profile
with a chart, and an anomalies sheet with conditional formatting.

Excel is the lingua franca of a trading desk, so the analysis ships in a form
an analyst or trader can open, filter, and build on directly.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill

from src.energyagent.analysis.timeseries import (
    detect_anomalies,
    hour_of_day_profile,
    peak_offpeak_summary,
)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def build_excel_report(
    df: pd.DataFrame,
    path: str | Path,
    metric_name: str = "Carbon intensity",
    unit: str = "gCO2/kWh",
    z: float = 2.0,
) -> Path:
    """Write a multi-sheet Excel report and return the file path."""
    path = Path(path)
    profile = hour_of_day_profile(df).rename(
        columns={"hour_uk": "Hour (UK)", "mean": f"Mean ({unit})",
                 "std": "Std dev", "count": "Samples"}
    )
    anomalies = detect_anomalies(df, z=z)
    # Excel can't store tz-aware datetimes: show them as UK local, tz-naive.
    if not anomalies.empty:
        from src.energyagent.analysis.timeseries import UK_TZ
        anomalies = anomalies.copy()
        anomalies["start_uk"] = (
            anomalies["start_utc"].dt.tz_convert(UK_TZ).dt.tz_localize(None)
        )
        anomalies = anomalies.drop(columns=["start_utc"])
        anomalies = anomalies[["start_uk"] + [c for c in anomalies.columns
                                              if c != "start_uk"]]
    summary = peak_offpeak_summary(df)
    summary_df = pd.DataFrame(
        [
            ["Metric", metric_name],
            ["Unit", unit],
            ["Readings analysed", int(len(df))],
            ["Cheapest/greenest hour (UK)", summary.get("best_hour")],
            [f"  its mean ({unit})", summary.get("best_mean")],
            ["Most expensive/dirtiest hour (UK)", summary.get("worst_hour")],
            [f"  its mean ({unit})", summary.get("worst_mean")],
            [f"Peak-to-trough spread ({unit})", summary.get("spread")],
            ["Anomalies flagged (|z| >= %.1f)" % z, int(len(anomalies))],
        ],
        columns=["Measure", "Value"],
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        profile.to_excel(writer, sheet_name="Hour profile", index=False)
        anomalies.to_excel(writer, sheet_name="Anomalies", index=False)

        wb = writer.book

        # Style headers on each sheet.
        _style_header(wb["Summary"], summary_df.shape[1])
        _style_header(wb["Hour profile"], profile.shape[1])
        if not anomalies.empty:
            _style_header(wb["Anomalies"], anomalies.shape[1])

        # Line chart of mean metric by hour of day.
        ws = wb["Hour profile"]
        chart = LineChart()
        chart.title = f"{metric_name} by hour of day"
        chart.y_axis.title = unit
        chart.x_axis.title = "Hour (UK)"
        nrows = profile.shape[0]
        data = Reference(ws, min_col=2, min_row=1, max_row=nrows + 1)  # Mean col
        cats = Reference(ws, min_col=1, min_row=2, max_row=nrows + 1)  # Hour col
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height, chart.width = 8, 16
        ws.add_chart(chart, "F2")

        # Conditional formatting: colour-scale the z-score column on Anomalies.
        if not anomalies.empty:
            wsa = wb["Anomalies"]
            zcol = anomalies.columns.get_loc("z_score") + 1
            last = anomalies.shape[0] + 1
            col_letter = wsa.cell(row=1, column=zcol).column_letter
            rng = f"{col_letter}2:{col_letter}{last}"
            wsa.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="num", start_value=-3, start_color="4472C4",
                    mid_type="num", mid_value=0, mid_color="FFFFFF",
                    end_type="num", end_value=3, end_color="C00000",
                ),
            )
    return path